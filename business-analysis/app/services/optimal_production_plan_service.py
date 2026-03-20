# -*- coding: utf-8 -*-
"""
最优生产方案服务
综合商业分析、绿色环保、智能检测，生成可一键导入设备/物料/工序/工艺的完整方案，
并导出 Excel、PDF。
"""
import io
import json
import os
from datetime import datetime
from typing import Optional

import httpx
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from app.models.schemas import (
    OptimalProductionPlanRequest,
    OptimalProductionPlanResult,
    PlanEquipmentItem,
    PlanMaterialItem,
    PlanProcedureItem,
    PlanProcessRouteItem,
    CostProfitData,
    MarketData,
)


def _default_equipment(product: str) -> list[PlanEquipmentItem]:
    """生成示例设备"""
    return [
        PlanEquipmentItem(equipment_code="EQ-SMT-01", equipment_name="SMT贴片机", manufacturer="", brand="", model="", supplier="", location="产线A", technical_params="精度0.1mm"),
        PlanEquipmentItem(equipment_code="EQ-AOI-01", equipment_name="AOI视觉检测", manufacturer="", brand="", model="", supplier="", location="产线A", technical_params="漏焊/短路检测"),
        PlanEquipmentItem(equipment_code="EQ-REFLOW-01", equipment_name="回流焊炉", manufacturer="", brand="", model="", supplier="", location="产线A", technical_params="8温区"),
    ]


def _default_materials(product: str) -> list[PlanMaterialItem]:
    """生成示例物料"""
    return [
        PlanMaterialItem(material_code="MAT-PCB-001", material_name="PCB基板", model="", quantity=1, unit="件", supplier=""),
        PlanMaterialItem(material_code="MAT-IC-001", material_name="主控芯片", model="", quantity=1, unit="颗", supplier=""),
        PlanMaterialItem(material_code="MAT-SOLDER-001", material_name="锡膏", model="", quantity=0.05, unit="g", supplier=""),
    ]


def _default_procedures(product: str) -> list[PlanProcedureItem]:
    """生成示例工序"""
    return [
        PlanProcedureItem(procedure_code="PROC-01", procedure_name="锡膏印刷", production_steps="钢网印刷", production_inspection_equipment="", operator="", equipment_codes=["EQ-SMT-01"], material_codes=["MAT-SOLDER-001"]),
        PlanProcedureItem(procedure_code="PROC-02", procedure_name="SMT贴片", production_steps="贴片焊接", production_inspection_equipment="", operator="", equipment_codes=["EQ-SMT-01"], material_codes=["MAT-PCB-001", "MAT-IC-001"]),
        PlanProcedureItem(procedure_code="PROC-03", procedure_name="回流焊接", production_steps="8温区回流", production_inspection_equipment="", operator="", equipment_codes=["EQ-REFLOW-01"], material_codes=[]),
        PlanProcedureItem(procedure_code="PROC-04", procedure_name="AOI质检", production_steps="视觉检测漏焊/短路/断路", production_inspection_equipment="AOI", operator="", equipment_codes=["EQ-AOI-01"], material_codes=[]),
    ]


def _default_process_routes(product: str, proc_codes: list[str]) -> list[PlanProcessRouteItem]:
    """生成示例工艺路线"""
    return [
        PlanProcessRouteItem(
            route_code="ROUTE-001",
            route_name=f"{product}工艺路线",
            product=product,
            version="V1.0",
            description="锡膏印刷→贴片→回流→AOI质检",
            procedure_sequence=proc_codes,
        )
    ]


def _build_flowchart_mermaid(equipment: list, materials: list, procedures: list, routes: list) -> str:
    """生成 Mermaid 流程图"""
    lines = ["flowchart TB", "  subgraph 设备"]
    for i, e in enumerate(equipment):
        name = (e.equipment_name or "").replace("[", "(").replace("]", ")").replace('"', "'")
        lines.append(f'    E{i}["{name}"]')
    lines.append("  end")
    lines.append("")
    lines.append("  subgraph 物料")
    for i, m in enumerate(materials):
        name = (m.material_name or "").replace("[", "(").replace("]", ")").replace('"', "'")
        lines.append(f'    M{i}["{name}"]')
    lines.append("  end")
    lines.append("")
    lines.append("  subgraph 工序")
    for i, p in enumerate(procedures):
        name = (p.procedure_name or "").replace("[", "(").replace("]", ")").replace('"', "'")
        lines.append(f'    P{i}["{name}"]')
    lines.append("  end")
    lines.append("")
    # 工序顺序
    if routes and routes[0].procedure_sequence:
        seq = routes[0].procedure_sequence
        proc_code_to_idx = {p.procedure_code: i for i, p in enumerate(procedures)}
        for j in range(len(seq) - 1):
            a = proc_code_to_idx.get(seq[j])
            b = proc_code_to_idx.get(seq[j + 1])
            if a is not None and b is not None:
                lines.append(f"  P{a} --> P{b}")
    else:
        for i in range(len(procedures) - 1):
            lines.append(f"  P{i} --> P{i+1}")
    return "\n".join(lines)


async def _call_llm(system: str, user: str, api_key: str, base_url: str, model: str) -> str:
    url = f"{base_url.rstrip('/')}/chat/completions"
    payload = {"model": model, "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}], "temperature": 0.5, "max_tokens": 1024}
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(url, json=payload, headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"})
        resp.raise_for_status()
        data = resp.json()
    choices = data.get("choices", [])
    if not choices:
        return ""
    return choices[0].get("message", {}).get("content", "").strip()


async def generate_optimal_production_plan(
    req: OptimalProductionPlanRequest,
    *,
    api_key: Optional[str] = None,
    base_url: Optional[str] = None,
    model: Optional[str] = None,
) -> OptimalProductionPlanResult:
    """
    综合商业分析、绿色环保、智能检测，生成最优生产方案。
    设备/物料/工序/工艺为空时自动生成示例数据，便于一键导入。
    """
    api_key = api_key or os.getenv("LLM_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("DASHSCOPE_API_KEY")
    base_url = (base_url or os.getenv("LLM_BASE_URL") or "https://dashscope.aliyuncs.com/compatible-mode/v1").rstrip("/")
    model = model or os.getenv("LLM_MODEL", "qwen-plus")

    equipment = req.equipment if req.equipment else _default_equipment(req.product_name)
    materials = req.materials if req.materials else _default_materials(req.product_name)
    procedures = req.procedures if req.procedures else _default_procedures(req.product_name)
    proc_codes = [p.procedure_code for p in procedures]
    process_routes = req.process_routes if req.process_routes else _default_process_routes(req.product_name, proc_codes)
    if process_routes and not process_routes[0].procedure_sequence:
        process_routes[0].procedure_sequence = proc_codes

    flowchart_mermaid = _build_flowchart_mermaid(equipment, materials, procedures, process_routes)

    # 商业分析摘要
    business_analysis = {}
    if req.cost_profit:
        cp = req.cost_profit
        direct = cp.bom_cost + cp.labor_cost + cp.energy_cost
        overhead = direct * cp.overhead_rate
        total_cost = direct + overhead + cp.logistics_delay_cost
        revenue = cp.selling_price * cp.quantity
        profit = revenue - total_cost
        business_analysis = {
            "bom_cost": cp.bom_cost,
            "labor_cost": cp.labor_cost,
            "energy_cost": cp.energy_cost,
            "total_cost": total_cost,
            "revenue": revenue,
            "profit": profit,
            "profit_margin": (profit / revenue * 100) if revenue > 0 else 0,
        }
    if req.market_data:
        business_analysis["target_industry"] = req.market_data.target_industry
        business_analysis["user_role"] = req.market_data.user_profile.role

    # 绿色环保摘要
    green_metrics = req.carbon_footprint or {}
    if req.green_supply_chain:
        green_metrics["supply_chain_scores"] = req.green_supply_chain

    # 智能检测配置
    inspection_config = req.defect_config or {}
    if req.anomaly_config:
        inspection_config["anomaly_detection"] = req.anomaly_config

    # AI 综合摘要
    summary = ""
    if api_key:
        lines = [
            f"产品：{req.product_name}，计划产量：{req.quantity}",
            f"设备数：{len(equipment)}，物料数：{len(materials)}，工序数：{len(procedures)}，工艺路线数：{len(process_routes)}",
        ]
        if business_analysis:
            lines.append(f"商业分析：总成本 {business_analysis.get('total_cost', 0):.2f} 元，利润 {business_analysis.get('profit', 0):.2f} 元")
        if green_metrics:
            lines.append("绿色环保：已纳入碳足迹与供应链评分")
        lines.append("智能检测：AOI 视觉质检，次品浪费折算，异常检测预防")
        lines.append("")
        lines.append("请用 2-3 句话输出「最优生产方案综合评估摘要」，突出成本、环保、质量三方面平衡。")
        try:
            summary = await _call_llm(
                "你是工业制造方案专家，擅长综合评估生产方案。",
                "\n".join(lines),
                api_key, base_url, model,
            )
        except Exception:
            pass

    if not summary:
        summary = (
            f"本方案涵盖 {req.product_name} 的完整生产流程："
            f"设备 {len(equipment)} 台、物料 {len(materials)} 种、工序 {len(procedures)} 道、工艺路线 {len(process_routes)} 条。"
            "综合商业分析、绿色环保与智能检测，可实现成本可控、低碳生产、零缺陷目标。"
        )

    return OptimalProductionPlanResult(
        success=True,
        product_name=req.product_name,
        quantity=req.quantity,
        summary=summary,
        business_analysis=business_analysis,
        green_metrics=green_metrics,
        inspection_config=inspection_config,
        equipment=equipment,
        materials=materials,
        procedures=procedures,
        process_routes=process_routes,
        flowchart_mermaid=flowchart_mermaid,
    )


def export_production_plan_to_excel(plan: OptimalProductionPlanResult) -> bytes:
    """导出生产方案为 Excel（含设备、物料、工序、工艺、商业分析、绿色环保、智能检测、流程图）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "导入说明"
    ws0.append(["产品生产方案 - 一键导入说明"])
    ws0.append([])
    ws0.append(["Sheet 名称", "导入目标", "说明"])
    ws0.append(["设备", "设备管理", "设备编码、设备名称、生产厂家、品牌、规格型号、供应商、位置、技术参数"])
    ws0.append(["物料", "物料管理", "物料编号、物料名称、规格型号、用量、单位、供应商"])
    ws0.append(["工序", "工序管理", "工序编号、工序名称、生产步骤、检测设备、操作人员、关联设备编码、关联物料编码"])
    ws0.append(["工艺路线", "工艺管理", "工艺编号、工艺名称、所属产品、版本、工艺描述、工序顺序"])
    ws0.append(["工艺路线-工序", "工艺管理", "route_id、procedure_id、sequence_order（需导入工序后获取ID）"])
    ws0.append([])
    ws0.append(["商业分析", "-", "利润、成本、市场策略摘要"])
    ws0.append(["绿色环保", "-", "碳足迹、供应链评分摘要"])
    ws0.append(["智能检测", "-", "质检配置、浪费折算、异常检测"])
    ws0.append(["流程图", "-", "Mermaid 代码，可复制到 Mermaid 编辑器渲染"])

    # 设备（表头与前端批量导入 CSV 列配置一致，便于一键导入）
    ws_eq = wb.create_sheet("设备", 1)
    ws_eq.append(["设备编码", "设备名称", "生产厂家", "品牌", "规格型号", "供应商", "生产日期", "使用年限", "折旧方式", "位置", "技术参数", "备品备件"])
    for e in plan.equipment:
        ws_eq.append([e.equipment_code, e.equipment_name, e.manufacturer, e.brand, e.model, e.supplier, "", "", "PrintinUsingTime", e.location, e.technical_params, ""])
    # 物料（表头与物料管理导入一致）
    ws_mat = wb.create_sheet("物料", 2)
    ws_mat.append(["物料编号", "物料名称", "规格型号", "库存数量", "供应商"])
    for m in plan.materials:
        ws_mat.append([m.material_code, m.material_name, m.model, m.quantity, m.supplier])

    # 工序（表头与工序管理导入一致）
    ws_proc = wb.create_sheet("工序", 3)
    ws_proc.append(["工序编号", "工序名称", "生产步骤", "生产和检测设备", "操作人员", "开始时间", "结束时间", "工序描述"])
    for p in plan.procedures:
        ws_proc.append([p.procedure_code, p.procedure_name, p.production_steps, p.production_inspection_equipment, p.operator, "", "", ""])

    # 工艺路线（表头与工艺管理导入一致）
    ws_route = wb.create_sheet("工艺路线", 4)
    ws_route.append(["工艺编号", "工艺名称", "所属产品", "版本", "工艺描述", "操作人员", "设备使用情况", "工序顺序"])
    for r in plan.process_routes:
        ws_route.append([r.route_code, r.route_name, r.product, r.version, r.description, "", "", ",".join(r.procedure_sequence)])

    # 商业分析
    ws_biz = wb.create_sheet("商业分析", 5)
    ws_biz.append(["项目", "数值/说明"])
    ws_biz.append(["产品", plan.product_name])
    ws_biz.append(["计划产量", plan.quantity])
    ws_biz.append(["综合摘要", plan.summary])
    for k, v in plan.business_analysis.items():
        ws_biz.append([str(k), str(v)])

    # 绿色环保
    ws_green = wb.create_sheet("绿色环保", 6)
    ws_green.append(["项目", "数值/说明"])
    for k, v in plan.green_metrics.items():
        val = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)
        ws_green.append([str(k), val])

    # 智能检测
    ws_insp = wb.create_sheet("智能检测", 7)
    ws_insp.append(["项目", "数值/说明"])
    for k, v in plan.inspection_config.items():
        val = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)
        ws_insp.append([str(k), val])

    # 流程图
    ws_flow = wb.create_sheet("流程图", 8)
    ws_flow.append(["Mermaid 流程图代码"])
    for line in plan.flowchart_mermaid.split("\n"):
        ws_flow.append([line])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_production_plan_to_pdf(plan: OptimalProductionPlanResult) -> bytes:
    """导出生产方案为 PDF（含上述所有信息）"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="PlanTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=10)
    h2_style = ParagraphStyle(name="PlanH2", parent=styles["Heading2"], fontSize=12, spaceAfter=6)

    story = []
    story.append(Paragraph(f"产品生产方案 - {plan.product_name}", title_style))
    story.append(Paragraph(f"生成日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("一、综合评估摘要", h2_style))
    story.append(Paragraph(plan.summary, styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("二、商业分析", h2_style))
    if plan.business_analysis:
        rows = [["项目", "数值"]]
        for k, v in plan.business_analysis.items():
            rows.append([str(k), str(v)])
        t = Table(rows, colWidths=[4*cm, 8*cm])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
        story.append(t)
    else:
        story.append(Paragraph("无", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("三、绿色环保", h2_style))
    if plan.green_metrics:
        for k, v in plan.green_metrics.items():
            story.append(Paragraph(f"{k}: {json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v}", styles["Normal"]))
    else:
        story.append(Paragraph("无", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("四、智能检测", h2_style))
    if plan.inspection_config:
        for k, v in plan.inspection_config.items():
            story.append(Paragraph(f"{k}: {json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v}", styles["Normal"]))
    else:
        story.append(Paragraph("AOI 视觉质检、次品浪费折算、异常检测预防", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("五、设备清单", h2_style))
    rows = [["设备编码", "设备名称", "规格型号", "位置"]]
    for e in plan.equipment:
        rows.append([e.equipment_code, e.equipment_name, e.model, e.location])
    t = Table(rows, colWidths=[3*cm, 4*cm, 4*cm, 3*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("六、物料清单", h2_style))
    rows = [["物料编号", "物料名称", "用量", "单位"]]
    for m in plan.materials:
        rows.append([m.material_code, m.material_name, str(m.quantity), m.unit])
    t = Table(rows, colWidths=[3*cm, 5*cm, 3*cm, 2*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("七、工序清单", h2_style))
    rows = [["工序编号", "工序名称", "生产步骤"]]
    for p in plan.procedures:
        rows.append([p.procedure_code, p.procedure_name, p.production_steps])
    t = Table(rows, colWidths=[3*cm, 4*cm, 7*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("八、工艺路线", h2_style))
    rows = [["工艺编号", "工艺名称", "工序顺序"]]
    for r in plan.process_routes:
        rows.append([r.route_code, r.route_name, " → ".join(r.procedure_sequence)])
    t = Table(rows, colWidths=[3*cm, 5*cm, 6*cm])
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.grey), ("GRID", (0, 0), (-1, -1), 0.5, colors.black)]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("九、流程图（Mermaid）", h2_style))
    story.append(Paragraph(plan.flowchart_mermaid.replace("\n", "<br/>"), styles["Normal"]))

    doc.build(story)
    return buffer.getvalue()
